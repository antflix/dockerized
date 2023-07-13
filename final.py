import datetime
import json
import sys as sys
from typing import Any
from flask import (Flask, Response, app, send_file, render_template,
                   request)
from openpyxl import load_workbook
import requests
from duplex import calc_quad, calc_duplex, accumulated_materials, reset_variables, calc_decora, calc_gfci, calc_cutin, calc_surface, calc_duplex_controlled, calc_quad, calc_quad_decora, calc_quad_gfci, calc_quad_cutin, calc_quad_surface, calc_quad_controlled, calc_ff3, calc_ff4, calc_rough_data, calc_cutin_data, calc_lv_switch, calc_hv_switch, calc_hv_dimming, calc_wh_120, calc_wh_277, calc_ff3
import os.path
from google.oauth2 import service_account

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from collections import OrderedDict
from google.oauth2 import service_account
from googleapiclient.discovery import build
import json
from openpyxl import Workbook

import gspread


# DEBUG = True
SITE_NAME = "Materialized"

items = {}

app = Flask(__name__)
@app.route('/count.html', methods=['GET', 'POST'])
def count():
    if request.method == 'POST':
        item = request.form['item']
        if item not in items:
            items[item] = 0
    return render_template('count.html', items=items)


@app.route('/calendar.html')
def calendar():
    # Pass the cell values to the template
    return render_template('calendar.html')


@app.route('/increment', methods=['POST'])
def increment():
    item = request.form['item']
    items[item] += 1
    return str(items[item])
  

@app.route('/index.html', methods=['GET', 'POST'])
@app.route('/', methods=['GET', 'POST'], )  # type: ignore
def form() -> Any:
    # global decora, duplex_outlets, gfci, cutins, surface, quad_standard, duplex_controlled, four_square_box, four_square_bracket_box
    # If User hits the Generate button...
    # If Generate button is not pressed, just show index.html
    global form_data

    if request.method == 'POST':
        reset_variables()
        reset_dict()
        duplex_outlets = (request.form['standard'])
        if duplex_outlets:
            duplex_outlets = int(duplex_outlets)
            form_data['duplex_outlets'] = int(request.form.get('standard', 0))
        else:
            duplex_outlets  = 0  
            form_data['duplex_outlets'] = 0

                        
        decora = (request.form['decora'])
        if decora:
            decora = int(decora)
            form_data['decora'] = int(request.form.get('decora', 0))
        else:
            decora  = 0  
            form_data['decora'] = 0
            
        gfci = (request.form['gfci'])
        if gfci:
            gfci = int(gfci)
            form_data['gfci'] = int(request.form.get('gfci', 0))
        else:
            gfci  = 0
            
        cutin = (request.form['cutin'])
        if cutin:
            cutin = int(cutin)
            form_data['cutin'] = int(request.form.get('cutin', 0))
        else:
            cutin  = 0
            
        surface = (request.form['surface'])
        if surface:
            surface = int(surface)  
            form_data['surface'] = int(request.form.get('surface', 0))
        else:
            surface  = 0
            
        duplex_controlled = (request.form['1switch'])
        if duplex_controlled:
            duplex_controlled = int(duplex_controlled)
            form_data['duplex_controlled'] = int(request.form.get('1switch', 0))
        else:
            duplex_controlled  = 0
        
        quad_standard = (request.form['quad_standard'])
        if quad_standard:
            quad_standard = int(quad_standard)
            form_data['quad_standard'] = int(request.form.get('quad_standard', 0))
        else:
            quad_standard  = 0  
    
        quad_decora = (request.form['quad_decora'])
        if quad_decora:
            quad_decora = int(quad_decora)
            form_data['quad_decora'] = int(request.form.get('quad_decora', 0))
        else:
            quad_decora  = 0
            
        quad_gfci = (request.form['quad_gfci'])
        if quad_gfci:
            quad_gfci = int(quad_gfci)
            form_data['quad_gfci'] = int(request.form.get('quad_gfci', 0))
        else:
            quad_gfci  = 0
            
        quad_cutin = (request.form['quad_cutin'])
        if quad_cutin:
            quad_cutin = int(quad_cutin)
            form_data['quad_cutin'] = int(request.form.get('quad_cutin', 0))
        else:
            quad_cutin  = 0
            
        quad_surface = (request.form['quad_surface'])
        if quad_surface:
            quad_surface = int(quad_surface)
            form_data['quad_surface'] = int(request.form.get('quad_surface', 0))
        else:
            quad_surface  = 0
            
        quad_controlled = (request.form['quad_controlled'])
        if quad_controlled:
            quad_controlled = int(quad_controlled)
            form_data['quad_controlled'] = int(request.form.get('quad_controlled', 0))
        else:
            quad_controlled  = 0

        ff3 = (request.form['3-wire'])
        if ff3:
            ff3 = int(ff3)
            form_data['ff3'] = int(request.form.get('3-wire', 0))
        else:
            ff3  = 0
            
        ff4 = (request.form['4-wire'])
        if ff4:
            ff4 = int(ff4)
            form_data['ff4'] = int(request.form.get('4-wire', 0))
        else:
            ff4  = 0

        rough_data = (request.form['rough_in_data'])
        if rough_data:
            rough_data = int(rough_data)
            form_data['rough_data'] = int(request.form.get('rough_in_data', 0))
        else:
            rough_data  = 0
            
        cutin_data = (request.form['cutin_data'])
        if cutin_data:
            cutin_data = int(cutin_data)
            form_data['cutin_data'] = int(request.form.get('cutin_data', 0))
        else:
            cutin_data  = 0

        lv_switch = (request.form['lv_switch'])
        if lv_switch:
            lv_switch = int(lv_switch)
            form_data['lv_switch'] = int(request.form.get('lv_switch', 0))
        else:
            lv_switch  = 0
            
        hv_switch = (request.form['hv_switch'])
        if hv_switch:
            hv_switch = int(hv_switch)
            form_data['hv_switch'] = int(request.form.get('hv_switch', 0))
        else:
            hv_switch  = 0
            
        hv_dimming = (request.form['hv_dimming'])
        if hv_dimming:
            hv_dimming = int(hv_dimming)
            form_data['hv_dimming'] = int(request.form.get('hv_dimming', 0))
        else:
            hv_dimming  = 0

        wh_120 = (request.form['wh_120'])
        if wh_120:
            wh_120 = int(wh_120)
            form_data['wh_120'] = int(request.form.get('wh_120', 0))
        else:
            wh_120  = 0

        wh_277 = (request.form['wh_277'])
        if wh_277:
            wh_277 = int(wh_277)
            form_data['wh_277'] = int(request.form.get('wh_277', 0))
        else:
            wh_277  = 0
        workbook = Workbook()
        sheet = workbook.active

        # Write form data to the worksheet
        
        sheet['E5'] = duplex_outlets
        sheet['E6'] = decora
        sheet['E7'] = gfci
        sheet['E8'] = cutin
        sheet['E9'] = surface
        sheet['E10'] = duplex_controlled
        
        sheet['E13'] = quad_standard
        sheet['E14'] = quad_decora
        sheet['E15'] = quad_gfci                                                                                                                                                                
        sheet['E16'] = quad_cutin
        sheet['E17'] = quad_surface
        sheet['E18'] = quad_controlled 
                                        
        sheet['E21'] = lv_switch
        sheet['E22'] = hv_switch
        sheet['E23'] = hv_dimming
              
        sheet['E26'] = rough_data
        sheet['E27'] = cutin_data                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        sheet['E28'] = ff3
        sheet['E29'] = ff4
        sheet['E30'] = wh_120
        sheet['E31'] = wh_277
        
        # need to get rid of standard/decora counts and add a toggle so that the total plug count can include cut-in and surface mounted with the correct type of outlet
        sheet['C5'] = "Standard"
        sheet['C6'] = "Decora" 
        sheet['C7'] = "GFCI"
        sheet['C8'] = "Cut-in"
        sheet['C9'] = "Surface Mounted"
        sheet['C10'] = "Controlled"
        # need to add sub catagory of single outlet vs whole outlet controlled(subtract amount of controlled from whole plug count)
        
        # need to get rid of standard/decora counts and refer to toggle above so that the total plug count can include cut-in and surface mounted with the correct type of outlet
        sheet['C13'] = "Standard"
        sheet['C14'] = "Decora"
        sheet['C15'] = "GFCI"                                                                                                                                                               
        sheet['C16'] = "Cut-in"
        sheet['C17'] = "Surface Mounted"
        sheet['C18'] = "Controlled"
        # need to add sub catagory of single outlet vs whole outlet controlled(subtract amount of controlled from whole plug count)

                          
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
        sheet['C21'] = "Low-Voltage"
        sheet['C22'] = "Line-Voltage"
        sheet['C23'] = "Line-Volt Dimming"
        
        sheet['C26'] = "Rough-in Data"
        sheet['C27'] = "Cut-in Data"
        sheet['C28'] = "3-wire Furniture Feed"
        sheet['C29'] = "4-wire Funiture Feed"
        sheet['C30'] = "208V 40A Instahot"
        sheet['C31'] = "277V 30A Instahot"


#  Save the workbook as an XLSX file
        workbook.save('form_data.xlsx')
        calc_duplex(duplex_outlets) 
        calc_decora(decora)
        calc_gfci(gfci)
        calc_cutin(cutin)
        calc_surface(surface)
        calc_duplex_controlled(duplex_controlled)
    
        calc_quad(quad_standard)
        calc_quad_decora(quad_decora)
        calc_quad_gfci(quad_gfci)
        calc_quad_cutin(quad_cutin)
        calc_quad_surface(quad_surface)
        calc_quad_controlled(quad_controlled)
        
        calc_ff3(ff3)
        calc_ff4(ff4)
        calc_rough_data(rough_data)
        calc_cutin_data(cutin_data)
        calc_lv_switch(lv_switch)
        calc_hv_switch(hv_switch)
        calc_hv_dimming(hv_dimming)
        calc_wh_120(wh_120)
        calc_wh_277(wh_277) 
    
        return render_template('result.html', materials=accumulated_materials, form_data=form_data), reset_variables()
    
        #Redirect to result page
        # return redirect(url_for('result'))  # type: ignore
    # show html page
    return render_template('index.html')  # type: ignore

def reset_dict():
    global form_data
    form_data = {}
    
@app.route('/download') # type: ignore
def donload() -> Response:
    filename = 'templates/new.xlsx'
    return send_file(filename, as_attachment=True)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
    



@app.route('/todo.html')
def todo() -> str:   # show the form, it wasn't submitted
    return render_template('todo.html')
@app.route('/upload-to-sheets', methods=['POST'])
def upload_to_sheets():
    # Google Sheets API credentials
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    spreadsheet_id = '1vCEfGo4MA-fwOtQAnPa6U69uzzjFGjlKzTJWNpUBoT4'
    range_xlsx = '# of Installed Devices!A1:Z100'  # Range for XLSX data
    range_html = 'Walls Material List!C12:F100'  # Range for HTML table data
    secret_file = os.path.join(os.getcwd(), 'credentials.json')

    # Load the XLSX file
    xlsx_file_path = "form_data.xlsx"
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    # Get all the values from the sheet
    data_xlsx = sheet.iter_rows(values_only=True)
    # Assuming the first row contains column headers, skip it
    headers_xlsx = next(data_xlsx)
    rows_xlsx = list(data_xlsx)

    # Convert datetime objects to strings
    rows_xlsx = [[str(cell) if isinstance(cell, datetime.datetime) else cell for cell in row] for row in rows_xlsx]

    # Parse the HTML table data from the POST request
    data_html = json.loads(request.data)
    headers_html = data_html[0]
    rows_html = data_html[1:]

    # Swap the sides of columns in HTML table data
    rows_html_swapped = [[row[i] for i in range(len(row)-1, -1, -1)] for row in rows_html]
    # Authenticate with Google Sheets API
    creds = service_account.Credentials.from_service_account_file(secret_file, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)

    try:
        # Prepare the XLSX data for uploading to Google Sheets
        values_xlsx = {'values': [headers_xlsx] + rows_xlsx}

        # Call the Google Sheets API to update the spreadsheet with the XLSX data
        result_xlsx = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_xlsx,
            valueInputOption='USER_ENTERED', body=values_xlsx
        ).execute()

        print(f"XLSX data uploaded to Google Sheets: {result_xlsx.get('updatedCells')} cells updated")

        # Prepare the HTML table data with swapped columns for uploading to Google Sheets
        values_html = {'values': rows_html_swapped}

        # Call the Google Sheets API to update the spreadsheet with the HTML table data
        result_html = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_html,
            valueInputOption='USER_ENTERED', body=values_html
        ).execute()

        print(f"HTML table data uploaded to Google Sheets: {result_html.get('updatedCells')} cells updated")

        return "Data uploaded to Google Sheets successfully!"
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return "Error occurred while uploading data to Google Sheets"


@app.route('/result.html')
def result():
    # Pass the cell values to the template
    return render_template('result.html')
    # return   send_file(filename, as_attachment=True, ) # type: ignore


@app.route('/manifest.json')  # type: ignore
def manifest() -> Response:
    return app.send_static_file('manifest.json')

# Serve service worker file


@app.route('/sw.js')
def service_worker() -> Response:
    return app.send_static_file('sw.js')

# Cache static assets


@app.after_request
def add_header(response) -> Response:
    response.headers['Cache-Control'] = 'static, max-age=31536000'
    return response


if __name__ == '__main__':
    app.run(debug = False, port = 5050, host = '0.0.0.0')
